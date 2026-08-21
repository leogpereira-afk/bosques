#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Extrai GerenciadorLoteamento.xlsx → seed/*.json (fora do git: tem CPF).

Uso:  python3 scripts/extrair-planilha.py [caminho do xlsx]

Saída em seed/:
  lotes.json       — 1 por lote com preço (as linhas-modelo vazias ficam de fora)
  clientes.json    — CPF normalizado (só dígitos) é a chave
  corretores.json  — idem
  vendas.json      — venda ligada a lote/cliente/corretor + plano da parcela
  despesas.json    — caixa, saídas
  receitas.json    — caixa, entradas avulsas ("Outras Receitas")
  pendencias.json  — tudo que NÃO deu para casar sozinho (vai para conferência)

Regras que vêm das fórmulas da planilha (conferidas em 21/08/2026, batendo
o Totalizador NO CENTAVO):
  • Valor da parcela da venda = Lotes!K (fixa c/ desconto) ou Lotes!M
    (reajustada 6% c/ desconto), conforme o plano diz "Fixa" ou "Reajustada".
  • 1ª parcela = 1 mês após DT RECEB ENTRADA; fim = início + qtde parcelas.
  • Em Despesas/Outras Receitas, SÓ A COLUNA C conta no caixa (é o que o
    Totalizador soma). Valor escrito na coluna E é ANOTAÇÃO de dinheiro já
    contado em outra aba (comissão que está em Vendas!O, entrada que está em
    Vendas!H). Esses saem com anotacao:true e ficam FORA das somas.
  • "Parcelas" do Totalizador é PROJEÇÃO (parcela × contratos ativos no mês),
    não recebimento: a planilha nunca controlou parcela a parcela.
"""
import json, re, sys, unicodedata
from datetime import datetime, date
from pathlib import Path

import openpyxl

RAIZ = Path(__file__).resolve().parent.parent
XLSX = Path(sys.argv[1]) if len(sys.argv) > 1 else Path.home() / 'Downloads/GerenciadorLoteamento.xlsx'
SAIDA = RAIZ / 'seed'
SAIDA.mkdir(exist_ok=True)

wb = openpyxl.load_workbook(XLSX, data_only=True)
pendencias = []


def digitos(v):
    return re.sub(r'\D', '', str(v or ''))


def texto(v):
    if v is None:
        return ''
    return re.sub(r'\s+', ' ', str(v)).strip()


def dinheiro(v):
    """Número ou texto '1.234,56' → float; senão None."""
    if v is None or v == '':
        return None
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    t = str(v).strip().replace('R$', '').strip()
    t = t.replace('.', '').replace(',', '.') if ',' in t else t
    try:
        return round(float(t), 2)
    except ValueError:
        return None


def dia(v):
    if isinstance(v, (datetime, date)):
        return v.strftime('%Y-%m-%d')
    m = re.match(r'^\s*(\d{1,2})/(\d{1,2})/(\d{4})\s*$', str(v or ''))
    if m:  # data digitada como texto ("27/05/2026") — o Totalizador a perde; nós não
        return f'{m.group(3)}-{int(m.group(2)):02d}-{int(m.group(1)):02d}'
    return None


# ---- Lotes -----------------------------------------------------------------
lotes = []
for r in wb['Lotes'].iter_rows(min_row=2, values_only=True):
    cod, quadra, num = r[0], r[1], r[2]
    if cod is None or r[8] in (None, 0):
        continue  # linha-modelo sem preço não é lote real
    lotes.append({
        'id': f'q{int(quadra)}-l{int(num)}',
        'cod': int(cod), 'quadra': int(quadra), 'lote': int(num),
        'rua': texto(r[3]), 'numero': texto(r[4]),
        'areaM2': dinheiro(r[5]), 'frente': dinheiro(r[6]), 'comprimento': dinheiro(r[7]),
        'preco': dinheiro(r[8]),
        # os 4 preços de parcela que a planilha carrega por lote
        'parcFixa': dinheiro(r[9]), 'parcFixaDesc': dinheiro(r[10]),
        'parcReaj': dinheiro(r[11]), 'parcReajDesc': dinheiro(r[12]),
        'status': 'Disponível',  # vendas.json muda depois, no import
    })
porCod = {l['cod']: l for l in lotes}

# ---- Clientes ----------------------------------------------------------------
clientes = {}
for r in wb['Clientes'].iter_rows(min_row=2, values_only=True):
    cpf = digitos(r[0])
    nome = texto(r[1])
    if not nome:
        continue
    if not cpf:
        pendencias.append({'onde': 'Clientes', 'quem': nome, 'problema': 'sem CPF'})
    clientes[cpf or nome] = {
        'cpf': cpf, 'nome': nome,
        'endereco': texto(r[2]), 'bairro': texto(r[3]), 'cidade': texto(r[4]),
        'uf': texto(r[5]), 'cep': texto(r[6]),
        'celular': texto(r[7]), 'whatsapp': texto(r[8]), 'email': texto(r[9]),
    }

# ---- Corretores ---------------------------------------------------------------
corretores = {}
for r in wb['Corretores'].iter_rows(min_row=2, values_only=True):
    cpf = digitos(r[0])
    nome = texto(r[1])
    if not nome:
        continue
    if cpf in corretores:  # Raphael aparece 2x (com e sem máscara no CPF)
        continue
    corretores[cpf or nome] = {
        'cpf': cpf, 'nome': nome, 'cidade': texto(r[4]), 'uf': texto(r[5]),
        'celular': texto(r[7]), 'whatsapp': texto(r[8]), 'email': texto(r[9]),
        'banco': texto(r[10]), 'agencia': texto(r[11]), 'conta': texto(r[12]),
        'chavePix': texto(r[13]), 'obs': texto(r[14]) if len(r) > 14 else '',
    }

# ---- Planos de pagamento (Descrição → Fixa/Reajustada) -------------------------
planoTipo = {}
for r in wb['Planos de Pagamento'].iter_rows(min_row=2, values_only=True):
    desc = texto(r[0])
    if desc:
        planoTipo[desc] = {'entrada': dinheiro(r[1]), 'qtde': int(r[2] or 0), 'tipo': texto(r[3]) or 'Fixa'}

# ---- Vendas --------------------------------------------------------------------
vendas = []
for i, r in enumerate(wb['Vendas'].iter_rows(min_row=3, values_only=True), 3):
    cod = r[0]
    if cod is None:
        if any(isinstance(x, (int, float)) and x for x in r[7:10]):
            # venda órfã: tem entrada mas não tem lote (o Totalizador conta esse dinheiro!)
            pendencias.append({'onde': f'Vendas linha {i}', 'quem': texto(r[4]),
                               'problema': f'entrada de R$ {dinheiro(r[7])} SEM lote — decidir qual lote é (ou é sinal/reserva)',
                               'dado': {'cpf': digitos(r[3]), 'entrada': dinheiro(r[7]),
                                        'forma': texto(r[8]), 'data': dia(r[9])}})
        continue
    try:
        cod = int(cod)
    except (TypeError, ValueError):
        pendencias.append({'onde': f'Vendas linha {i}', 'problema': f'Cod Lote não numérico: {cod!r}'})
        continue
    lote = porCod.get(cod)
    if not lote:
        pendencias.append({'onde': f'Vendas linha {i}', 'problema': f'lote cod {cod} não existe na aba Lotes'})
        continue
    cpfCli = digitos(r[3])
    nomeCli = texto(r[4])
    if (cpfCli or nomeCli) and cpfCli not in clientes:
        # cliente da venda que não está na aba Clientes: cria pelo que a venda traz
        clientes[cpfCli or nomeCli] = {'cpf': cpfCli, 'nome': nomeCli, 'endereco': '', 'bairro': '',
                                       'cidade': '', 'uf': '', 'cep': '', 'celular': '', 'whatsapp': '', 'email': ''}
        pendencias.append({'onde': f'Vendas linha {i}', 'quem': nomeCli,
                           'problema': 'cliente só existia na venda — cadastro criado incompleto'})
    plano = texto(r[6])
    info = planoTipo.get(plano, {})
    tipoParc = info.get('tipo', '')
    if plano and not info:
        # o nome do plano na venda não bate 1:1 com a aba de planos → deduz pelo texto
        tipoParc = 'Reajustada' if re.search(r'reajust|6\s*%', plano, re.I) else 'Fixa'
        pendencias.append({'onde': f'Vendas linha {i}', 'problema': f'plano "{plano[:60]}" fora da aba Planos — tipo deduzido: {tipoParc}'})
    qtde = int(r[10] or 0)
    vendas.append({
        'loteId': lote['id'], 'loteCod': cod, 'quadra': lote['quadra'], 'lote': lote['lote'],
        'clienteCpf': cpfCli, 'clienteNome': nomeCli,
        'dataVenda': dia(r[5]) or dia(r[9]),   # DATA DA VENDA quase sempre vazia → cai na entrada
        'plano': plano, 'tipoParcela': tipoParc or 'Fixa',
        'entrada': dinheiro(r[7]), 'formaEntrada': texto(r[8]), 'dataEntrada': dia(r[9]),
        'qtdeParcelas': qtde,
        'valorParcela': dinheiro(r[11]) if r[11] not in (None, '') else
                        (lote['parcFixaDesc'] if (tipoParc or 'Fixa') == 'Fixa' else lote['parcReajDesc']),
        'corretorCpf': digitos(r[12]), 'corretorNome': texto(r[13]),
        'comissao': dinheiro(r[14]), 'comissaoData': dia(r[15]), 'comissaoForma': texto(r[16]),
        'inicioParcelas': dia(r[17]), 'fimParcelas': dia(r[18]),
        'obs': ' | '.join(t for t in (texto(x) for x in r[19:24]) if t),
    })

# vendas duplicadas do mesmo lote?
vistos = {}
for v in vendas:
    if v['loteId'] in vistos:
        pendencias.append({'onde': 'Vendas', 'problema':
                           f"lote {v['loteId']} vendido 2x: {vistos[v['loteId']]} e {v['clienteNome']}"})
    vistos[v['loteId']] = v['clienteNome']

# ---- Despesas / Outras Receitas -------------------------------------------------
def lancamentos(aba):
    """Coluna C = conta no caixa. Valor só na coluna E = anotação de dinheiro
    já contado em outra aba (sai com anotacao:true, fora das somas)."""
    out = []
    for i, r in enumerate(wb[aba].iter_rows(min_row=2, values_only=True), 2):
        d, desc = dia(r[0]), texto(r[1])
        if not desc:
            continue
        valor = dinheiro(r[2])
        anotacao = False
        forma = texto(r[3])
        if valor is None:
            valor = dinheiro(r[4]) if len(r) > 4 else None
            anotacao = valor is not None
            obs = texto(r[5]) if len(r) > 5 else ''
        else:
            obs = ' '.join(t for t in (texto(x) for x in r[4:6]) if t) if len(r) > 4 else ''
        if valor is None:
            pendencias.append({'onde': f'{aba} linha {i}', 'quem': desc, 'problema': 'sem valor legível'})
            continue
        if not d:
            pendencias.append({'onde': f'{aba} linha {i}', 'quem': desc,
                               'problema': f'sem data (valor R$ {valor}) — entra no caixa só depois de datar'})
        out.append({'data': d, 'descricao': desc, 'valor': valor, 'forma': forma,
                    'anotacao': anotacao, 'obs': obs})
    return out

despesas = lancamentos('Despesas')
receitas = lancamentos('Outras Receitas')

# ---- grava -----------------------------------------------------------------------
def grava(nome, dado):
    p = SAIDA / f'{nome}.json'
    p.write_text(json.dumps(dado, ensure_ascii=False, indent=1), encoding='utf-8')
    n = len(dado)
    print(f'  {nome}.json — {n}')

print(f'Extraído de {XLSX.name}:')
grava('lotes', lotes)
grava('clientes', list(clientes.values()))
grava('corretores', list(corretores.values()))
grava('vendas', vendas)
grava('despesas', despesas)
grava('receitas', receitas)
grava('pendencias', pendencias)
def em2026(d):
    return bool(d) and d.startswith('2026')

soma_d = sum(x['valor'] for x in despesas if not x['anotacao'])
soma_r = sum(x['valor'] for x in receitas if not x['anotacao'])
soma_e = sum(v['entrada'] or 0 for v in vendas if em2026(v['dataEntrada']))
soma_c = sum(v['comissao'] or 0 for v in vendas if em2026(v['comissaoData']))
print('\nConferência contra o Totalizador da planilha (2026):')
print(f'  entradas recebidas em 2026 : R$ {soma_e:,.2f}  (Totalizador: 217.868,87 — ele soma +1.000 da venda órfã sem lote)')
print(f'  comissões pagas em 2026    : R$ {soma_c:,.2f}  (Totalizador: 108.314,20)')
print(f'  despesas (caixa)           : R$ {soma_d:,.2f}  (Totalizador: 227.179,46 — ele perde 5.001,10 de 3 linhas com data quebrada, que NÓS guardamos)')
print(f'  outras receitas (caixa)    : R$ {soma_r:,.2f}  (Totalizador: 123.834,60)')
anot_d = sum(x['valor'] for x in despesas if x['anotacao'])
anot_r = sum(x['valor'] for x in receitas if x['anotacao'])
print(f'  anotações fora do caixa    : despesas R$ {anot_d:,.2f} | receitas R$ {anot_r:,.2f} (dinheiro já contado em Vendas)')
